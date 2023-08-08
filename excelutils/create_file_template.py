from excelutils import ExcelControl, headers, width_columns, formats
from openpyxl.styles import (NamedStyle, Font, Border, Side, PatternFill, Alignment)
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet import worksheet


# https://stackoverflow.com/questions/27133731/folding-multiple-rows-with-openpyxl
#
def styles_add(this_book: Workbook):
    name_styles = {}
    for item in formats:
        name_styles[item['name']] = NamedStyle(name=item['name'])
        # header_style = NamedStyle(name=item['name'])
        name_styles[item['name']].font = Font(name=item['font']['name'], bold=item['font']['bold'],
                                              size=item['font']['size'])
        bd = Side(style=item['border']['style'], color=item['border']['color'])
        name_styles[item['name']].border = Border(left=bd, top=bd, right=bd, bottom=bd)
        name_styles[item['name']].fill = PatternFill(patternType=item['fill']['patternType'],
                                                     fgColor=item['fill']['fgColor'])
        name_styles[item['name']].alignment = Alignment(horizontal=item['alignment']['horizontal'],
                                                        vertical=item['alignment']['vertical'],
                                                        wrap_text=item['alignment']['wrap_text'],
                                                        shrink_to_fit=item['alignment']['shrink_to_fit'],
                                                        indent=item['alignment']['indent'])
        if not (item['name'] in this_book.named_styles):
            this_book.add_named_style(name_styles[item['name']])


def create_basic_header(operate_file: ExcelControl, sheet_name: str):
    sheet = operate_file.book[sheet_name]
    sheet.append(["."])
    sheet.append(headers["A2"])
    for column in range(1, len(headers["A2"]) + 1):
        sheet.cell(row=2, column=column).style = 'title_basic'
    sheet.cell(row=1, column=operate_file.column_number['K']).value = headers['K1']
    sheet.cell(row=1, column=operate_file.column_number['K']).style = 'title_basic'
    sheet.merge_cells('K1:M1')

    sheet.cell(row=1, column=operate_file.column_number['N']).value = headers['N1']
    sheet.cell(row=1, column=operate_file.column_number['N']).style = 'title_basic'
    sheet.merge_cells('N1:O1')

    sheet.cell(row=1, column=operate_file.column_number['P']).value = headers['P1']
    sheet.cell(row=1, column=operate_file.column_number['P']).style = 'title_basic'
    sheet.merge_cells('P1:S1')

    for width in width_columns:
        sheet.column_dimensions[width].width = width_columns[width]


def prepare_file_template(file_name: str = None, file_path: str = None, sheets_name: list[str] = None):
    output_file = ExcelControl(file_name, file_path)
    with output_file as ex:
        ex.delete_all_sheets()
        ex.create_sheets(sheets_name)
        print(output_file)
        print(type(ex.book))
        styles_add(ex.book)
        create_basic_header(ex, 'name')


if __name__ == "__main__":
    fln = "template_x_xx.xlsx"
    sheets = ["name", "stat"]
    prepare_file_template(fln, "../output", sheets)
