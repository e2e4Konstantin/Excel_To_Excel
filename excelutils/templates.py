from openpyxl.styles import (NamedStyle, Font, Border, Side, PatternFill, Alignment)
from openpyxl.workbook.workbook import Workbook
from openpyxl.utils.cell import column_index_from_string
from openpyxl.worksheet import worksheet
import re
from excelutils import headers, width_columns, formats
from common_data import QuoteInfo


# https://stackoverflow.com/questions/27133731/folding-multiple-rows-with-openpyxl
#
def styles_add(this_book: Workbook):
    name_styles = {}
    for item in formats:
        name_styles[item['name']] = NamedStyle(name=item['name'])
        # header_style = NamedStyle(name=item['name'])
        name_styles[item['name']].font = Font(name=item['font']['name'], bold=item['font']['bold'],
                                              size=item['font']['size'])
        bd = Side(style=item['border']['style'], color=item['border']['color'])
        name_styles[item['name']].border = Border(left=bd, top=bd, right=bd, bottom=bd)
        name_styles[item['name']].fill = PatternFill(patternType=item['fill']['patternType'],
                                                     fgColor=item['fill']['fgColor'])
        name_styles[item['name']].alignment = Alignment(horizontal=item['alignment']['horizontal'],
                                                        vertical=item['alignment']['vertical'],
                                                        wrap_text=item['alignment']['wrap_text'],
                                                        shrink_to_fit=item['alignment']['shrink_to_fit'],
                                                        indent=item['alignment']['indent'])
        if not (item['name'] in this_book.named_styles):
            this_book.add_named_style(name_styles[item['name']])


def create_basic_header(sheet: worksheet):
    sheet.append(["."])
    headers["A:O"].extend(headers["P:T"])
    sheet.append(headers["A:O"])
    for column in range(1, len(headers["A:O"]) + 1):
        sheet.cell(row=2, column=column).style = 'title_basic'
    sheet.cell(row=1, column=column_index_from_string('K')).value = headers['K1']
    sheet.cell(row=1, column=column_index_from_string('K')).style = 'title_basic'
    sheet.merge_cells('K1:M1')

    sheet.cell(row=1, column=column_index_from_string('N')).value = headers['N1']
    sheet.cell(row=1, column=column_index_from_string('N')).style = 'title_basic'
    sheet.merge_cells('N1:O1')

    sheet.cell(row=1, column=column_index_from_string('P')).value = headers['P1']
    sheet.cell(row=1, column=column_index_from_string('P')).style = 'title_basic'
    sheet.merge_cells('P1:T1')

    for width in width_columns:
        sheet.column_dimensions[width].width = width_columns[width]


def range_decorating(sheet: worksheet, row, columns: list[str], style_name: str):
    """ Устанавливает стиль style_name для ячеек из списка columns """
    for column in columns:
        sheet.cell(row=row, column=column_index_from_string(column)).style = style_name


def create_table_header(sheet: worksheet, table_info: tuple[str, str, str, str], row: int):
    """ На лист sheet в строку row, информацию table_info """
    tmp = re.findall(r'(^\d+\.\d+)(-\d+)+(-\d+$)', table_info[0])[0]
    extended_name = f"Таблица {tmp[0]}{tmp[2]}. {table_info[1].capitalize()}."

    sheet.cell(row=row, column=column_index_from_string('E')).value = table_info[0]  # код таблицы
    sheet.cell(row=row, column=column_index_from_string('F')).value = ""
    sheet.cell(row=row, column=column_index_from_string('G')).value = extended_name  # table_info[1]  # название таблицы
    sheet.cell(row=row, column=column_index_from_string('H')).value = ""
    sheet.cell(row=row, column=column_index_from_string('I')).value = ""
    sheet.cell(row=row, column=column_index_from_string('J')).value = ""

    range_decorating(sheet, row, ['E', 'F', 'H', 'I', 'J', ], 'line_table')
    sheet.cell(row=row, column=column_index_from_string('G')).style = 'table_name'

    column = column_index_from_string('K')
    sheet.cell(row=row - 1, column=column).value = headers['K1']
    sheet.cell(row=row - 1, column=column).style = 'further_quotes'
    sheet.merge_cells(start_row=row - 1, start_column=column, end_row=row - 1, end_column=column + 2)

    sheet.cell(row=row, column=column_index_from_string('K')).value = headers['K']
    sheet.cell(row=row, column=column_index_from_string('L')).value = headers['L']
    sheet.cell(row=row, column=column_index_from_string('M')).value = headers['M']
    range_decorating(sheet, row, ['K', 'L', 'M'], 'further_quotes')

    column = column_index_from_string('N')
    sheet.cell(row=row - 1, column=column).value = headers['N1']
    sheet.cell(row=row - 1, column=column).style = 'title_attributes'

    attributes = []
    if table_info[2]:
        attributes.extend(table_info[2].split(','))  # атрибуты
        attributes_length = len(attributes)
        for i, attribute in enumerate(attributes):
            sheet.cell(row=row, column=column + i).value = attribute
            sheet.cell(row=row, column=column + i).style = 'title_attributes'

        if attributes_length > 1:
            sheet.merge_cells(start_row=row - 1, start_column=column, end_row=row - 1,
                              end_column=column + attributes_length - 1)
    else:
        sheet.cell(row=row, column=column_index_from_string('N')).value = headers['N']
        sheet.cell(row=row, column=column_index_from_string('O')).value = headers['O']

        sheet.cell(row=row, column=column_index_from_string('N')).style = 'title_attributes'
        sheet.cell(row=row, column=column_index_from_string('O')).style = 'title_attributes'

        sheet.merge_cells(start_row=row - 1, start_column=column, end_row=row - 1, end_column=column + 1)

    parameters = []
    if table_info[3]:
        column += len(attributes)
        parameters.extend(table_info[3].split(','))  # параметры
        parameters_length = len(parameters)
        if parameters_length > 0:
            mini_table_header = headers['P:T']
            parameter_tile_length = len(mini_table_header)  # таблица: 'от', 'до', 'ед.изм.', 'шаг', 'тип'

            for parameter in parameters:
                sheet.cell(row=row - 1, column=column).value = parameter  # название параметра
                sheet.cell(row=row - 1, column=column).style = 'title_parameter'

                i = 0
                for item in mini_table_header:
                    sheet.cell(row=row, column=column + i).value = item
                    sheet.cell(row=row, column=column + i).style = 'title_parameter'
                    i += 1

                column_delta = parameter_tile_length - 1
                sheet.merge_cells(start_row=row - 1, start_column=column, end_row=row - 1,
                                  end_column=column + column_delta)
                column += column_delta + 1
    # sheet.append(["."])


def put_table_to_sheet(sheet: worksheet, table_info: tuple[str, str, str, str], row: int):
    """ На лист sheet в строку row, пишет информацию о таблице table_info.
     Присваивает группу """
    sheet.row_dimensions.group(row, row, outline_level=1)
    create_table_header(sheet, table_info, row)


def create_quote_line(sheet: worksheet, quote_info: QuoteInfo, row: int):
    """ На лист sheet в строку row, информацию table_info """
    print(row, quote_info)
    sheet.cell(row=row, column=column_index_from_string('E')).value = quote_info[0]  # код таблицы
    sheet.cell(row=row, column=column_index_from_string('F')).value = quote_info[1]
    sheet.cell(row=row, column=column_index_from_string('G')).value = quote_info[2]
    sheet.cell(row=row, column=column_index_from_string('H')).value = quote_info[3]
    sheet.cell(row=row, column=column_index_from_string('I')).value = int(quote_info[4]) if quote_info[4].isdigit() else ""
    sheet.cell(row=row, column=column_index_from_string('J')).value = '++'

    range_decorating(sheet, row, ['E', 'F', 'G', 'H', 'I', 'J'], 'quote_line')
    sheet.cell(row=row, column=column_index_from_string('I')).font = Font(name='Calibri', bold=False, size=8, color="000000")
    sheet.cell(row=row, column=column_index_from_string('I')).alignment = Alignment(horizontal='right')
    sheet.cell(row=row, column=column_index_from_string('J')).alignment = Alignment(horizontal='center')

def put_quote_to_sheet(sheet: worksheet, quote_info: QuoteInfo, row: int):
    """ На лист sheet в строку row, пишет информацию о расценке quote_info.
     Присваивает группу """
    sheet.row_dimensions.group(row, row, outline_level=2)
    create_quote_line(sheet, quote_info, row)
